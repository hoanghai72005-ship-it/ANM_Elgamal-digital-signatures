import sys
import json
import math
import random
import threading
import os

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QGroupBox, QLabel, QLineEdit, QPushButton,
    QRadioButton, QTabWidget, QTextEdit, QMessageBox, QFileDialog,
    QButtonGroup, QDialog, QSpinBox, QComboBox
)
from PySide6.QtCore import Qt, Signal, QObject
from PySide6.QtGui import QClipboard, QFont, QTextCharFormat

import elgamal_core as core

STYLE_SHEET = """
    /* TOOLTIP */
    QToolTip {
        background-color: #2b2b2b;
        color: #ffffff;
        border: 1px solid #767676;
        padding: 6px;
        border-radius: 4px;
        font-size: 13px;
    }

    QRadioButton::indicator {
        width: 14px;
        height: 14px;
        border-radius: 9px; 
        border: 2px solid #000000; 
        background-color: #ffffff; 
    }
    QRadioButton::indicator:checked {
        border: 2px solid #0078D4; 
        background-color: #ffffff; 
    }
    QRadioButton::indicator:hover {
        border: 2px solid #666666;
    }
    QPushButton#NutBoGoc {
        background-color: #0078D4; 
        color: white;
        border-radius: 15px; 
        padding: 8px 15px;
        font-weight: bold;
        border: none;
    }
    QPushButton#NutBoGoc:hover {
        background-color: #005A9E; 
    }
    QPushButton#NutBoGoc:disabled {
        background-color: #CCE4F7; 
        color: #ffffff;
    }
    QPushButton {
        padding: 6px 12px;
        border-radius: 3px;
        border: 1px solid #ccc;
        background-color: #f9f9f9;
    }
    QPushButton:hover {
        background-color: #e9e9e9;
    }
"""


class SoanThaoDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("✍️ Trình Soạn Thảo Văn Bản Cần Ký")
        self.resize(650, 450)
        layout = QVBoxLayout(self)

        toolbar = QHBoxLayout()
        self.btn_bold = QPushButton("B")
        self.btn_bold.setCheckable(True)
        self.btn_bold.setStyleSheet("font-weight: bold;")
        self.btn_bold.clicked.connect(self.toggle_bold)

        self.btn_italic = QPushButton("I")
        self.btn_italic.setCheckable(True)
        self.btn_italic.setStyleSheet("font-style: italic; font-family: serif;")
        self.btn_italic.clicked.connect(self.toggle_italic)

        self.spin_size = QSpinBox()
        self.spin_size.setRange(8, 72)
        self.spin_size.setValue(12)
        self.spin_size.valueChanged.connect(self.change_font_size)

        toolbar.addWidget(self.btn_bold)
        toolbar.addWidget(self.btn_italic)
        toolbar.addWidget(QLabel(" Cỡ chữ: "))
        toolbar.addWidget(self.spin_size)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        self.editor = QTextEdit()
        layout.addWidget(self.editor)

        btn_layout = QHBoxLayout()
        btn_save = QPushButton("✔️ Lưu tệp & Sử dụng để Ký")
        btn_save.setObjectName("NutBoGoc")
        btn_save.clicked.connect(self.luu_va_dong)

        btn_cancel = QPushButton("❌ Hủy")
        btn_cancel.clicked.connect(self.reject)

        btn_layout.addStretch()
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_save)
        layout.addLayout(btn_layout)

        self.duong_dan_luu = None

    def toggle_bold(self):
        fmt = QTextCharFormat()
        fmt.setFontWeight(QFont.Bold if self.btn_bold.isChecked() else QFont.Normal)
        self.editor.mergeCurrentCharFormat(fmt)

    def toggle_italic(self):
        fmt = QTextCharFormat()
        fmt.setFontItalic(self.btn_italic.isChecked())
        self.editor.mergeCurrentCharFormat(fmt)

    def change_font_size(self, size):
        fmt = QTextCharFormat()
        fmt.setFontPointSize(size)
        self.editor.mergeCurrentCharFormat(fmt)

    def luu_va_dong(self):
        if not self.editor.toPlainText().strip():
            QMessageBox.warning(self, "Cảnh báo", "Văn bản đang trống, không có gì để ký!")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Lưu văn bản vào máy để thực hiện Ký", "van_ban_soan_thao.html",
                                              "HTML Files (*.html);;Text Files (*.txt)")
        if path:
            try:
                with open(path, 'w', encoding='utf-8') as f:
                    if path.endswith('.html'):
                        f.write(self.editor.toHtml())
                    else:
                        f.write(self.editor.toPlainText())
                self.duong_dan_luu = path
                self.accept()
            except Exception as e:
                QMessageBox.critical(self, "Lỗi", f"Không thể lưu file: {e}")


class WorkerSignals(QObject):
    hoan_thanh_sinh_khoa = Signal(str, str, str)
    hoan_thanh_ma_hoa = Signal(str)
    hoan_thanh_giai_ma = Signal(str)
    hoan_thanh_ky = Signal(str)
    hoan_thanh_xac_thuc = Signal(str)
    bao_loi = Signal(str, str, str)


class UngDungElGamal(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Hệ mật mã ElGamal - Chữ ký số ElGamal")
        self.resize(1100, 750)
        self.setStyleSheet(STYLE_SHEET)

        self.q = self.a = self.XA = self.YA = None
        self.k_ma_hoa_da_chot = None
        self.k_chu_ky_da_chot = None
        self.tin_hieu = WorkerSignals()

        self.tin_hieu.hoan_thanh_sinh_khoa.connect(self._phan_hoi_giao_dien_sinh_khoa)
        self.tin_hieu.hoan_thanh_ma_hoa.connect(self._phan_hoi_giao_dien_ma_hoa)
        self.tin_hieu.hoan_thanh_giai_ma.connect(self._phan_hoi_giao_dien_giai_ma)
        self.tin_hieu.hoan_thanh_ky.connect(self._phan_hoi_giao_dien_ky)
        self.tin_hieu.hoan_thanh_xac_thuc.connect(self._phan_hoi_xac_thuc_chi_tiet_ui)
        self.tin_hieu.bao_loi.connect(self._hien_thi_loi_tu_luong)

        widget_chinh = QWidget()
        self.setCentralWidget(widget_chinh)
        layout_chinh = QVBoxLayout(widget_chinh)

        self.tao_khung_sinh_khoa(layout_chinh)
        self.tao_cac_the(layout_chinh)

        self.chuyen_doi_che_do_khoa()
        self.chuyen_doi_k_ma_hoa()
        self.chuyen_doi_k_chu_ky()

    def rut_gon_chuoi_hien_thi(self, chuoi, max_len=50):
        chuoi_str = str(chuoi)
        if len(chuoi_str) <= max_len:
            return chuoi_str

        nua_chieu_dai = (max_len - 5) // 2
        return chuoi_str[:nua_chieu_dai] + " ... " + chuoi_str[-nua_chieu_dai:]

    # 1. GIAO DIỆN
    def tao_khung_sinh_khoa(self, layout_cha):
        khung_khoa = QGroupBox("🔑 Giai đoạn sinh khóa")
        layout = QGridLayout(khung_khoa)

        self.radio_khoa_tuy_chon = QRadioButton("Tùy chọn (Nhập tay)")
        self.radio_khoa_tu_dong = QRadioButton("Tự động chọn")
        self.radio_khoa_tu_dong.setChecked(True)
        self.nhom_khoa = QButtonGroup()
        self.nhom_khoa.addButton(self.radio_khoa_tuy_chon)
        self.nhom_khoa.addButton(self.radio_khoa_tu_dong)
        self.radio_khoa_tuy_chon.toggled.connect(self.chuyen_doi_che_do_khoa)

        layout.addWidget(self.radio_khoa_tuy_chon, 0, 0)
        layout.addWidget(self.radio_khoa_tu_dong, 0, 1)

        layout.addWidget(QLabel("Chọn số nguyên tố q đủ lớn:"), 1, 0)
        self.o_nhap_q = QLineEdit()
        self.o_nhap_q.setToolTip("💡 ĐIỀU KIỆN: q phải là số nguyên tố lớn (128/512/1024 bit).")
        layout.addWidget(self.o_nhap_q, 1, 1)

        layout.addWidget(QLabel("Chọn a là căn nguyên thủy của q:"), 2, 0)
        self.o_nhap_a = QLineEdit()
        self.o_nhap_a.setToolTip("💡 ĐIỀU KIỆN: 0 < a < q.")
        layout.addWidget(self.o_nhap_a, 2, 1)

        layout.addWidget(QLabel("Khóa bí mật XA:"), 3, 0)
        self.o_nhap_x = QLineEdit()
        self.o_nhap_x.setToolTip("💡 ĐIỀU KIỆN: Khóa bí mật phải là số nguyên dương XA, XA < q - 1.")
        layout.addWidget(self.o_nhap_x, 3, 1)

        khung_sinh_ngau_nhien = QWidget()
        lo_sinh_ngau_nhien = QHBoxLayout(khung_sinh_ngau_nhien)
        lo_sinh_ngau_nhien.setContentsMargins(0, 0, 0, 0)

        self.combo_bit = QComboBox()
        self.combo_bit.addItems(["128 bit", "512 bit", "1024 bit"])
        self.combo_bit.setCurrentIndex(0)

        self.nut_khoa_ngau_nhien = QPushButton("🎲 Tạo ngẫu nhiên")
        self.nut_khoa_ngau_nhien.clicked.connect(self.tao_khoa_ngau_nhien_ui)

        lo_sinh_ngau_nhien.addWidget(self.combo_bit)
        lo_sinh_ngau_nhien.addWidget(self.nut_khoa_ngau_nhien)
        layout.addWidget(khung_sinh_ngau_nhien, 0, 2)

        self.nut_xac_nhan_khoa = QPushButton("✔️ Tính YA và Xác nhận Khóa")
        self.nut_xac_nhan_khoa.setObjectName("NutBoGoc")
        self.nut_xac_nhan_khoa.clicked.connect(self.xac_nhan_khoa)
        layout.addWidget(self.nut_xac_nhan_khoa, 0, 3)

        # CẬP NHẬT: Thêm căn lề phải cho các nhãn Text (YA, Khóa công khai, Khóa bí mật)
        nhan_chu_ya = QLabel("YA = a^XA mod q =")
        nhan_chu_ya.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        layout.addWidget(nhan_chu_ya, 1, 2)

        self.nhan_ket_qua_ya = QLabel("")
        self.nhan_ket_qua_ya.setStyleSheet("font-weight: bold; color: #0078D4;")
        self.nhan_ket_qua_ya.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(self.nhan_ket_qua_ya, 1, 3)

        nhan_chu_kck = QLabel("Khóa công khai {q, a, YA}:")
        nhan_chu_kck.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        layout.addWidget(nhan_chu_kck, 2, 2)

        self.nhan_khoa_cong_khai = QLabel("")
        self.nhan_khoa_cong_khai.setStyleSheet("font-weight: bold; color: #0078D4;")
        self.nhan_khoa_cong_khai.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(self.nhan_khoa_cong_khai, 2, 3)

        nhan_chu_kbm = QLabel("Khóa bí mật {XA}:")
        nhan_chu_kbm.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        layout.addWidget(nhan_chu_kbm, 3, 2)

        self.nhan_khoa_bi_mat = QLabel("")
        self.nhan_khoa_bi_mat.setStyleSheet("font-weight: bold; color: #0078D4;")
        self.nhan_khoa_bi_mat.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(self.nhan_khoa_bi_mat, 3, 3)

        layout_cha.addWidget(khung_khoa)

    def tao_cac_the(self, layout_cha):
        self.so_tay = QTabWidget()

        the_ma_hoa = QWidget()
        self.tao_giao_dien_ma_hoa(the_ma_hoa)
        self.so_tay.addTab(the_ma_hoa, "🔏 Mã hóa Chuỗi Ký Tự")

        the_chu_ky = QWidget()
        self.tao_giao_dien_chu_ky(the_chu_ky)
        self.so_tay.addTab(the_chu_ky, "✍️ Chữ ký số ElGamal")

        layout_cha.addWidget(self.so_tay)

    def tao_giao_dien_ma_hoa(self, cha):
        layout = QHBoxLayout(cha)

        khung_trai = QGroupBox("Thực hiện Mã hóa")
        layout_trai = QVBoxLayout(khung_trai)

        layout_trai.addWidget(QLabel("Bản rõ:"))
        self.o_chu_ban_ro = QTextEdit()
        layout_trai.addWidget(self.o_chu_ban_ro)

        khung_cc_ro = QWidget()
        lo_cc_ro = QHBoxLayout(khung_cc_ro)
        lo_cc_ro.setContentsMargins(0, 0, 0, 0)
        btn_copy_ro = QPushButton("📋 Sao chép")
        btn_copy_ro.clicked.connect(lambda: self.hanh_dong_sao_chep(self.o_chu_ban_ro))
        btn_save_ro = QPushButton("💾 Tải File")
        btn_save_ro.clicked.connect(lambda: self.hanh_dong_luu_tep(self.o_chu_ban_ro, "ban_ro.txt"))
        lo_cc_ro.addStretch()
        lo_cc_ro.addWidget(btn_copy_ro)
        lo_cc_ro.addWidget(btn_save_ro)
        layout_trai.addWidget(khung_cc_ro)

        khung_k_ma_hoa = QWidget()
        lo_k_ma_hoa = QGridLayout(khung_k_ma_hoa)

        self.radio_k_mh_tuy = QRadioButton("Tùy chọn (Nhập tay)")
        self.radio_k_mh_tu_dong = QRadioButton("Tự động chọn")
        self.radio_k_mh_tu_dong.setChecked(True)
        self.nhom_k_mh = QButtonGroup()
        self.nhom_k_mh.addButton(self.radio_k_mh_tuy)
        self.nhom_k_mh.addButton(self.radio_k_mh_tu_dong)
        self.radio_k_mh_tuy.toggled.connect(self.chuyen_doi_k_ma_hoa)

        lo_k_ma_hoa.addWidget(self.radio_k_mh_tuy, 0, 0)
        lo_k_ma_hoa.addWidget(self.radio_k_mh_tu_dong, 0, 1)

        lo_k_ma_hoa.addWidget(QLabel("k ="), 1, 0)
        self.o_nhap_k_ma_hoa = QLineEdit()
        self.o_nhap_k_ma_hoa.setToolTip("💡 Số k là khóa phiên dùng một lần (session key)")
        lo_k_ma_hoa.addWidget(self.o_nhap_k_ma_hoa, 1, 1)

        self.nut_k_mh_ngau_nhien = QPushButton("🎲 Tạo ngẫu nhiên")
        self.nut_k_mh_ngau_nhien.clicked.connect(self.tao_k_ma_hoa_ngau_nhien_ui)
        lo_k_ma_hoa.addWidget(self.nut_k_mh_ngau_nhien, 1, 2)

        self.nut_xac_nhan_k_ma_hoa = QPushButton("✔️ Xác nhận k")
        self.nut_xac_nhan_k_ma_hoa.setObjectName("NutBoGoc")
        self.nut_xac_nhan_k_ma_hoa.clicked.connect(self.xac_nhan_k_ma_hoa)
        lo_k_ma_hoa.addWidget(self.nut_xac_nhan_k_ma_hoa, 1, 3)

        lo_k_ma_hoa.addWidget(QLabel("K = (YA^k) ="), 2, 0)
        self.o_nhap_K_lon_ma_hoa = QLineEdit()
        self.o_nhap_K_lon_ma_hoa.setReadOnly(True)
        lo_k_ma_hoa.addWidget(self.o_nhap_K_lon_ma_hoa, 2, 1)

        self.nut_ma_hoa = QPushButton("🔒 Thực hiện mã hóa")
        self.nut_ma_hoa.setObjectName("NutBoGoc")
        self.nut_ma_hoa.clicked.connect(self.hanh_dong_ma_hoa)
        self.nut_ma_hoa.setEnabled(False)
        lo_k_ma_hoa.addWidget(self.nut_ma_hoa, 2, 3)

        layout_trai.addWidget(khung_k_ma_hoa)

        layout_trai.addWidget(QLabel("Bản mã gửi đi:"))
        self.o_chu_ban_ma_gui = QTextEdit()
        layout_trai.addWidget(self.o_chu_ban_ma_gui)

        khung_cc_ma_hoa = QWidget()
        lo_cc_mh = QHBoxLayout(khung_cc_ma_hoa)
        lo_cc_mh.setContentsMargins(0, 0, 0, 0)
        btn_copy_mh = QPushButton("📋 Sao chép")
        btn_copy_mh.clicked.connect(lambda: self.hanh_dong_sao_chep(self.o_chu_ban_ma_gui))
        btn_save_mh = QPushButton("💾 Lưu File")
        btn_save_mh.clicked.connect(lambda: self.hanh_dong_luu_tep(self.o_chu_ban_ma_gui, "ban_ma_gui.txt"))
        lo_cc_mh.addStretch()
        lo_cc_mh.addWidget(btn_copy_mh)
        lo_cc_mh.addWidget(btn_save_mh)
        layout_trai.addWidget(khung_cc_ma_hoa)

        layout.addWidget(khung_trai)

        khung_phai = QGroupBox("Thực hiện giải mã")
        layout_phai = QVBoxLayout(khung_phai)

        layout_phai.addWidget(QLabel("Bản mã nhận được:"))
        self.o_chu_ban_ma_nhan = QTextEdit()
        layout_phai.addWidget(self.o_chu_ban_ma_nhan)

        self.nut_giai_ma = QPushButton("🔓 Thực hiện giải mã")
        self.nut_giai_ma.setObjectName("NutBoGoc")
        self.nut_giai_ma.setEnabled(False)
        self.nut_giai_ma.clicked.connect(self.hanh_dong_giai_ma)
        layout_phai.addWidget(self.nut_giai_ma)

        layout_phai.addWidget(QLabel("Bản được giải mã:"))
        self.o_chu_ban_giai_ma = QTextEdit()
        layout_phai.addWidget(self.o_chu_ban_giai_ma)

        khung_cc_giai_ma = QWidget()
        lo_cc = QHBoxLayout(khung_cc_giai_ma)
        lo_cc.setContentsMargins(0, 0, 0, 0)
        btn_copy_gm = QPushButton("📋 Sao chép")
        btn_copy_gm.clicked.connect(lambda: self.hanh_dong_sao_chep(self.o_chu_ban_giai_ma))
        btn_save_gm = QPushButton("💾 Lưu File")
        btn_save_gm.clicked.connect(lambda: self.hanh_dong_luu_tep(self.o_chu_ban_giai_ma, "ban_giai_ma.txt"))
        lo_cc.addStretch()
        lo_cc.addWidget(btn_copy_gm)
        lo_cc.addWidget(btn_save_gm)
        layout_phai.addWidget(khung_cc_giai_ma)

        layout.addWidget(khung_phai)

    def tao_giao_dien_chu_ky(self, cha):
        layout = QVBoxLayout(cha)

        khung_ky = QGroupBox("✍️ Thực hiện ký")
        layout_ky = QVBoxLayout(khung_ky)

        khung_k_ky = QWidget()
        lo_k_ky = QHBoxLayout(khung_k_ky)

        self.radio_k_ck_tuy = QRadioButton("Tùy chọn k")
        self.radio_k_ck_tu_dong = QRadioButton("Tự động chọn k")
        self.radio_k_ck_tu_dong.setChecked(True)
        self.nhom_k_ck = QButtonGroup()
        self.nhom_k_ck.addButton(self.radio_k_ck_tuy)
        self.nhom_k_ck.addButton(self.radio_k_ck_tu_dong)
        self.radio_k_ck_tuy.toggled.connect(self.chuyen_doi_k_chu_ky)

        lo_k_ky.addWidget(self.radio_k_ck_tuy)
        lo_k_ky.addWidget(self.radio_k_ck_tu_dong)

        lo_k_ky.addWidget(QLabel(" Số ngẫu nhiên k: "))
        self.o_nhap_k_chu_ky = QLineEdit()
        self.o_nhap_k_chu_ky.setToolTip(
            "💡 ĐIỀU KIỆN: 0 < k < q - 1 \ngcd(k,q-1) = 1")
        lo_k_ky.addWidget(self.o_nhap_k_chu_ky)

        self.nut_k_ck_ngau_nhien = QPushButton("🎲 Tạo ngẫu nhiên")
        self.nut_k_ck_ngau_nhien.clicked.connect(self.tao_k_chu_ky_ngau_nhien_ui)
        lo_k_ky.addWidget(self.nut_k_ck_ngau_nhien)

        self.nut_xac_nhan_k_chu_ky = QPushButton("✔️ Xác nhận k")
        self.nut_xac_nhan_k_chu_ky.setObjectName("NutBoGoc")
        self.nut_xac_nhan_k_chu_ky.clicked.connect(self.xac_nhan_k_chu_ky)
        lo_k_ky.addWidget(self.nut_xac_nhan_k_chu_ky)
        layout_ky.addWidget(khung_k_ky)

        khung_tep = QWidget()
        lo_tep = QHBoxLayout(khung_tep)
        lo_tep.addWidget(QLabel("Chọn file cần thực hiện ký:"))
        self.o_nhap_tep_ky = QLineEdit()
        lo_tep.addWidget(self.o_nhap_tep_ky)

        btn_chon_tep = QPushButton("📁 Tải file")
        btn_chon_tep.clicked.connect(self.chon_tep_de_ky)
        lo_tep.addWidget(btn_chon_tep)

        btn_soan_thao = QPushButton("✍️ Soạn thảo")
        btn_soan_thao.clicked.connect(self.mo_trinh_soan_thao)
        lo_tep.addWidget(btn_soan_thao)

        btn_xem_file = QPushButton("👁️ Xem nội dung")
        btn_xem_file.clicked.connect(lambda: self.xem_noi_dung_file(self.o_nhap_tep_ky.text().strip()))
        lo_tep.addWidget(btn_xem_file)

        self.nut_thuc_hien_ky = QPushButton("✍️ Thực hiện ký lên văn bản")
        self.nut_thuc_hien_ky.setObjectName("NutBoGoc")
        self.nut_thuc_hien_ky.setEnabled(False)
        self.nut_thuc_hien_ky.clicked.connect(self.hanh_dong_ky)
        lo_tep.addWidget(self.nut_thuc_hien_ky)
        layout_ky.addWidget(khung_tep)

        layout_ky.addWidget(QLabel("Tệp chữ ký được sinh ra (r, s):"))
        self.o_chu_chu_ky = QTextEdit()
        layout_ky.addWidget(self.o_chu_chu_ky)

        khung_cc_ky = QWidget()
        lo_cc_ky = QHBoxLayout(khung_cc_ky)
        lo_cc_ky.setContentsMargins(0, 0, 0, 0)
        btn_copy_ky = QPushButton("📋 Sao chép")
        btn_copy_ky.clicked.connect(lambda: self.hanh_dong_sao_chep(self.o_chu_chu_ky))
        btn_save_ky = QPushButton("💾 Lưu File JSON")
        btn_save_ky.clicked.connect(lambda: self.hanh_dong_luu_tep(self.o_chu_chu_ky, "chu_ky.json"))
        lo_cc_ky.addStretch()
        lo_cc_ky.addWidget(btn_copy_ky)
        lo_cc_ky.addWidget(btn_save_ky)
        layout_ky.addWidget(khung_cc_ky)

        layout.addWidget(khung_ky)

        khung_xt = QGroupBox("🛡️ Kiểm tra (Xác thực)")
        layout_xt = QVBoxLayout(khung_xt)

        khung_tep_xt = QWidget()
        lo_tep_xt = QHBoxLayout(khung_tep_xt)

        lo_tep_xt.addWidget(QLabel("Chọn file cần kiểm tra (Bản nhận được):"))

        self.o_nhap_tep_xac_thuc = QLineEdit()
        lo_tep_xt.addWidget(self.o_nhap_tep_xac_thuc)

        btn_chon_tep_xt = QPushButton("📁 Tải file")
        btn_chon_tep_xt.clicked.connect(self.chon_tep_de_xac_thuc)
        lo_tep_xt.addWidget(btn_chon_tep_xt)

        btn_xem_file_xt = QPushButton("👁️ Xem nội dung")
        btn_xem_file_xt.clicked.connect(lambda: self.xem_noi_dung_file(self.o_nhap_tep_xac_thuc.text().strip()))
        lo_tep_xt.addWidget(btn_xem_file_xt)

        layout_xt.addWidget(khung_tep_xt)

        khung_chu_ky_xt = QWidget()
        lo_chu_ky_xt = QHBoxLayout(khung_chu_ky_xt)
        lo_chu_ky_xt.setContentsMargins(0, 0, 0, 0)
        lo_chu_ky_xt.addWidget(QLabel("Nhập hoặc tải lên chuỗi chữ ký (r, s):"))

        btn_tai_chu_ky_xt = QPushButton("📁 Tải file chữ ký")
        btn_tai_chu_ky_xt.clicked.connect(self.chon_tep_chu_ky_de_xac_thuc)
        lo_chu_ky_xt.addStretch()
        lo_chu_ky_xt.addWidget(btn_tai_chu_ky_xt)

        layout_xt.addWidget(khung_chu_ky_xt)

        self.o_chu_chu_ky_xac_thuc = QTextEdit()
        layout_xt.addWidget(self.o_chu_chu_ky_xac_thuc)

        self.nut_thkh_xac_thuc = QPushButton("🛡️ Thực hiện Kiểm tra chữ ký")
        self.nut_thkh_xac_thuc.setObjectName("NutBoGoc")
        self.nut_thkh_xac_thuc.clicked.connect(self.hanh_dong_xac_thuc)
        layout_xt.addWidget(self.nut_thkh_xac_thuc)

        layout.addWidget(khung_xt)

    # 2. TIỆN ÍCH CHUNG
    def hanh_dong_sao_chep(self, text_edit):
        clipboard = QApplication.clipboard()
        clipboard.setText(text_edit.toPlainText())
        QMessageBox.information(self, "Thành công", "📋 Đã sao chép nội dung!")

    def hanh_dong_luu_tep(self, text_edit, ten_mac_dinh):
        noi_dung = text_edit.toPlainText().strip()
        if not noi_dung:
            return QMessageBox.warning(self, "Cảnh báo", "Không có nội dung để lưu!")
        duong_dan, _ = QFileDialog.getSaveFileName(self, "Lưu file", ten_mac_dinh,
                                                   "Text Document (*.txt);;JSON File (*.json);;All Files (*.*)")
        if duong_dan:
            try:
                with open(duong_dan, 'w', encoding='utf-8') as f:
                    f.write(noi_dung)
                QMessageBox.information(self, "Thành công", f"💾 Đã lưu thành công tại:\n{duong_dan}")
            except Exception as e:
                QMessageBox.critical(self, "Lỗi", f"Không thể lưu: {str(e)}")

    def mo_trinh_soan_thao(self):
        dlg = SoanThaoDialog(self)
        if dlg.exec() == QDialog.Accepted and dlg.duong_dan_luu:
            self.o_nhap_tep_ky.setText(dlg.duong_dan_luu)

    def xem_noi_dung_file(self, path):
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "Cảnh báo", "Đường dẫn file không hợp lệ hoặc chưa được chọn!")
            return

        ext = os.path.splitext(path)[1].lower()
        content = ""

        try:
            if ext in ['.txt', '.json', '.html']:
                with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
            elif ext in ['.doc', '.docx']:
                try:
                    import docx
                    doc = docx.Document(path)
                    content = "\n".join([p.text for p in doc.paragraphs])
                except ImportError:
                    content = "[LỖI HỆ THỐNG]: Vui lòng chạy lệnh 'pip install python-docx' để hỗ trợ giải mã file Word."
                except Exception as e:
                    content = f"[LỖI ĐỊNH DẠNG]: Không hỗ trợ hiển thị tệp .DOC phiên bản cũ. Chi tiết: {e}"
            elif ext == '.pdf':
                try:
                    import pypdf
                    reader = pypdf.PdfReader(path)
                    text_runs = []
                    for page in reader.pages:
                        extracted = page.extract_text()
                        if extracted:
                            text_runs.append(extracted)
                    content = "\n".join(text_runs) if text_runs else "[Tệp PDF không chứa văn bản Text thô]"
                except ImportError:
                    content = "[LỖI HỆ THỐNG]: Vui lòng chạy lệnh 'pip install pypdf' để hỗ trợ giải mã file PDF."
            elif ext in ['.xls', '.xlsx']:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, data_only=True)
                    lines = []
                    for sheet in wb.sheetnames:
                        lines.append(f"====== Trang tính (Sheet): {sheet} ======")
                        ws = wb[sheet]
                        for row in ws.iter_rows(values_only=True):
                            if any(row is not None for row in row):
                                line_cells = [str(cell) if cell is not None else "" for cell in row]
                                lines.append(" \t | \t ".join(line_cells))
                    content = "\n".join(lines)
                except ImportError:
                    content = "[LỖI HỆ THỐNG]: Vui lòng chạy lệnh 'pip install openpyxl' để hỗ trợ giải mã file Excel."
                except Exception as e:
                    content = f"[LỖI ĐỊNH DẠNG]: Lỗi hiển thị file Excel. Chi tiết: {e}"
            else:
                content = "[HỆ THỐNG]: Định dạng file lạ không hỗ trợ xem Text. Tuy nhiên thuật toán vẫn tính toán mã băm nhị phân bình thường!"

            dlg = QDialog(self)
            dlg.setWindowTitle(f"👁️ Nội dung chi tiết: {os.path.basename(path)}")
            dlg.resize(600, 450)
            l = QVBoxLayout(dlg)
            txt = QTextEdit()
            txt.setReadOnly(True)
            if path.endswith('.html'):
                txt.setHtml(content)
            else:
                txt.setPlainText(content)
            l.addWidget(txt)
            dlg.exec()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi phân tích file", f"Không thể trích xuất dữ liệu: {e}")

    def chon_tep_chu_ky_de_xac_thuc(self):
        file, _ = QFileDialog.getOpenFileName(
            self, "Chọn file chữ ký", "",
            "JSON Files (*.json);;Text Files (*.txt);;All Files (*.*)"
        )
        if file:
            try:
                with open(file, 'r', encoding='utf-8') as f:
                    noi_dung = f.read()

                try:
                    json.loads(noi_dung)
                except json.JSONDecodeError:
                    QMessageBox.warning(self, "Cảnh báo",
                                        "File được chọn dường như không phải là một chuỗi JSON hợp lệ. Hệ thống vẫn sẽ tải nội dung lên nhưng có thể gặp lỗi khi xác thực.")

                self.o_chu_chu_ky_xac_thuc.setText(noi_dung)
            except Exception as e:
                QMessageBox.critical(self, "Lỗi đọc file", f"Không thể tải file chữ ký: {e}")

    # 3. LOGIC SINH KHÓA TỐI ƯU
    def chuyen_doi_che_do_khoa(self):
        is_tuy_chon = self.radio_khoa_tuy_chon.isChecked()
        for widget in [self.o_nhap_q, self.o_nhap_a, self.o_nhap_x]:
            widget.setReadOnly(not is_tuy_chon)
            if is_tuy_chon: widget.clear()

        self.nut_khoa_ngau_nhien.setEnabled(not is_tuy_chon)
        self.combo_bit.setEnabled(not is_tuy_chon)

        if not is_tuy_chon:
            self.tao_khoa_ngau_nhien_ui()

    def tao_khoa_ngau_nhien_ui(self):
        chuoi_bit = self.combo_bit.currentText()
        do_dai_bit = int(chuoi_bit.split()[0])

        self.nut_khoa_ngau_nhien.setEnabled(False)
        self.setCursor(Qt.WaitCursor)

        def luong_sinh_khoa_ngam():
            try:
                q = core.sinh_so_nguyen_to_an_toan(do_dai_bit)
                while True:
                    a = random.randint(2, q - 1)
                    if core.la_can_nguyen_thuy(a, q):
                        break
                XA = random.randint(1, q - 2)
                self.tin_hieu.hoan_thanh_sinh_khoa.emit(str(q), str(a), str(XA))
            except Exception as e:
                self.tin_hieu.bao_loi.emit("Lỗi Sinh Khóa", str(e), "sinh_khoa")

        threading.Thread(target=luong_sinh_khoa_ngam, daemon=True).start()

    def _phan_hoi_giao_dien_sinh_khoa(self, q_str, a_str, XA_str):
        self.o_nhap_q.setText(q_str)
        self.o_nhap_a.setText(a_str)
        self.o_nhap_x.setText(XA_str)

        self.o_nhap_q.setCursorPosition(0)
        self.o_nhap_a.setCursorPosition(0)
        self.o_nhap_x.setCursorPosition(0)

        self.o_nhap_q.setToolTip(f"q = {q_str}")
        self.o_nhap_a.setToolTip(f"a = {a_str}")
        self.o_nhap_x.setToolTip(f"XA = {XA_str}")

        self.nut_khoa_ngau_nhien.setEnabled(True)
        self.setCursor(Qt.ArrowCursor)

    def xac_nhan_khoa(self):
        try:
            q = int(self.o_nhap_q.text())
            a = int(self.o_nhap_a.text())
            XA = int(self.o_nhap_x.text())

            YA = core.tao_khoa(q, a, XA)
            self.q, self.a, self.XA, self.YA = q, a, XA, YA

            chuoi_ya = str(YA)
            chuoi_kck = f"{{{q}, {a}, {YA}}}"
            chuoi_kbm = f"{{{XA}}}"

            self.nhan_ket_qua_ya.setText(self.rut_gon_chuoi_hien_thi(chuoi_ya, max_len=60))
            self.nhan_khoa_cong_khai.setText(self.rut_gon_chuoi_hien_thi(chuoi_kck, max_len=60))
            self.nhan_khoa_bi_mat.setText(self.rut_gon_chuoi_hien_thi(chuoi_kbm, max_len=60))

            self.nhan_ket_qua_ya.setToolTip(chuoi_ya)
            self.nhan_khoa_cong_khai.setToolTip(chuoi_kck)
            self.nhan_khoa_bi_mat.setToolTip(chuoi_kbm)

            for widget in [self.o_nhap_q, self.o_nhap_a, self.o_nhap_x]:
                widget.setReadOnly(True)

            self.chuyen_doi_k_ma_hoa()
            self.chuyen_doi_k_chu_ky()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi Xác Nhận", f"Dữ liệu khóa không hợp lệ: {e}")

    # 4. MÃ HÓA / GIẢI MÃ
    def chuyen_doi_k_ma_hoa(self):
        is_tuy_chon = self.radio_k_mh_tuy.isChecked()
        self.o_nhap_k_ma_hoa.setReadOnly(not is_tuy_chon)
        self.nut_k_mh_ngau_nhien.setEnabled(not is_tuy_chon)
        self.nut_ma_hoa.setEnabled(False)
        self.o_nhap_K_lon_ma_hoa.clear()
        if not is_tuy_chon:
            self.tao_k_ma_hoa_ngau_nhien_ui()
        else:
            self.o_nhap_k_ma_hoa.clear()

    def tao_k_ma_hoa_ngau_nhien_ui(self):
        if not self.q: return
        k_str = str(random.randint(1, self.q - 1))
        self.o_nhap_k_ma_hoa.setText(k_str)
        self.o_nhap_k_ma_hoa.setCursorPosition(0)
        self.o_nhap_k_ma_hoa.setToolTip(f"k = {k_str}")

    def xac_nhan_k_ma_hoa(self):
        if not self.q: return QMessageBox.warning(self, "Cảnh báo", "Vui lòng Xác nhận Khóa chung trước!")
        chuoi_k = self.o_nhap_k_ma_hoa.text().strip()
        if not chuoi_k.isdigit(): return QMessageBox.critical(self, "Lỗi", "k phải là số nguyên!")
        k = int(chuoi_k)
        if k <= 0 or k >= self.q: return QMessageBox.critical(self, "Lỗi", f"k phải thuộc (0, {self.q})")

        self.k_ma_hoa_da_chot = k
        self.o_nhap_k_ma_hoa.setReadOnly(True)
        Gia_tri_K = str(core.tinh_luy_thua_module(self.YA, k, self.q))

        self.o_nhap_K_lon_ma_hoa.setText(Gia_tri_K)
        self.o_nhap_K_lon_ma_hoa.setCursorPosition(0)
        self.o_nhap_K_lon_ma_hoa.setToolTip(f"K = {Gia_tri_K}")

        self.nut_ma_hoa.setEnabled(True)

    def hanh_dong_ma_hoa(self):
        ban_ro = self.o_chu_ban_ro.toPlainText().strip()
        if not ban_ro: return QMessageBox.warning(self, "Cảnh báo", "Vui lòng nhập bản rõ!")
        self.nut_ma_hoa.setEnabled(False)

        def run_thread():
            try:
                chuoi_b64 = core.ma_hoa_van_ban(self.q, self.a, self.YA, ban_ro, self.k_ma_hoa_da_chot)
                self.tin_hieu.hoan_thanh_ma_hoa.emit(chuoi_b64)
            except Exception as e:
                self.tin_hieu.bao_loi.emit("Lỗi Mã Hóa", str(e), "ma_hoa")

        threading.Thread(target=run_thread, daemon=True).start()

    def _phan_hoi_giao_dien_ma_hoa(self, chuoi_b64):
        self.o_chu_ban_ma_gui.setText(chuoi_b64)
        self.o_chu_ban_ma_nhan.setText(chuoi_b64)
        self.nut_ma_hoa.setEnabled(True)
        self.nut_giai_ma.setEnabled(True)

    def hanh_dong_giai_ma(self):
        chuoi_b64 = self.o_chu_ban_ma_nhan.toPlainText().strip()
        if not chuoi_b64: return
        self.nut_giai_ma.setEnabled(False)

        def run_thread():
            try:
                chuoi_giai_ma = core.giai_ma_van_ban(self.q, self.XA, chuoi_b64)
                self.tin_hieu.hoan_thanh_giai_ma.emit(chuoi_giai_ma)
            except Exception:
                self.tin_hieu.bao_loi.emit("Lỗi Giải Mã", "Bản mã không hợp lệ!", "giai_ma")

        threading.Thread(target=run_thread, daemon=True).start()

    def _phan_hoi_giao_dien_giai_ma(self, text):
        self.o_chu_ban_giai_ma.setText(text)
        self.nut_giai_ma.setEnabled(True)

    def _hien_thi_loi_tu_luong(self, tieu_de, noi_dung, tac_vu_bi_loi):
        self.setCursor(Qt.ArrowCursor)
        self.nut_khoa_ngau_nhien.setEnabled(True)
        QMessageBox.critical(self, tieu_de, noi_dung)
        if tac_vu_bi_loi == "ma_hoa":
            self.nut_ma_hoa.setEnabled(True)
        elif tac_vu_bi_loi == "giai_ma":
            self.nut_giai_ma.setEnabled(True)
        elif tac_vu_bi_loi == "ky":
            self.nut_thuc_hien_ky.setEnabled(True)
        elif tac_vu_bi_loi == "xac_thuc":
            self.nut_thkh_xac_thuc.setEnabled(True)

    # 5. CHỮ KÝ SỐ
    def chuyen_doi_k_chu_ky(self):
        is_tuy_chon = self.radio_k_ck_tuy.isChecked()
        self.o_nhap_k_chu_ky.setReadOnly(not is_tuy_chon)
        self.nut_k_ck_ngau_nhien.setEnabled(not is_tuy_chon)
        self.nut_thuc_hien_ky.setEnabled(False)
        if not is_tuy_chon:
            self.tao_k_chu_ky_ngau_nhien_ui()
        else:
            self.o_nhap_k_chu_ky.clear()

    def tao_k_chu_ky_ngau_nhien_ui(self):
        if not self.q: return
        while True:
            k = random.randint(1, self.q - 2)
            if math.gcd(k, self.q - 1) == 1:
                k_str = str(k)
                self.o_nhap_k_chu_ky.setText(k_str)
                self.o_nhap_k_chu_ky.setCursorPosition(0)
                self.o_nhap_k_chu_ky.setToolTip(f"k = {k_str}")
                break

    def xac_nhan_k_chu_ky(self):
        if not self.q: return QMessageBox.warning(self, "Cảnh báo", "Xác nhận Khóa trước!")
        chuoi_k = self.o_nhap_k_chu_ky.text().strip()
        if not chuoi_k.isdigit(): return QMessageBox.critical(self, "Lỗi", "k phải là số nguyên!")
        k = int(chuoi_k)
        if math.gcd(k, self.q - 1) != 1:
            return QMessageBox.critical(self, "Lỗi", "gcd(k, q-1) phải bằng 1!")
        self.k_chu_ky_da_chot = k
        self.o_nhap_k_chu_ky.setReadOnly(True)
        self.nut_thuc_hien_ky.setEnabled(True)

    def chon_tep_de_ky(self):
        file, _ = QFileDialog.getOpenFileName(
            self, "Chọn tài liệu thực hiện ký số Elgamal", "",
            "Tài liệu Elgamal (*.txt *.json *.doc *.docx *.pdf *.xls *.xlsx);;All Files (*.*)"
        )
        if file: self.o_nhap_tep_ky.setText(file)

    def chon_tep_de_xac_thuc(self):
        file, _ = QFileDialog.getOpenFileName(
            self, "Chọn tài liệu cần kiểm tra chữ ký", "",
            "Tài liệu Elgamal (*.txt *.json *.doc *.docx *.pdf *.xls *.xlsx);;All Files (*.*)"
        )
        if file: self.o_nhap_tep_xac_thuc.setText(file)

    def hanh_dong_ky(self):
        duong_dan = self.o_nhap_tep_ky.text().strip()
        if not duong_dan: return
        self.nut_thuc_hien_ky.setEnabled(False)

        def run_thread():
            try:
                r, s, m = core.ky_tep(self.q, self.a, self.XA, self.k_chu_ky_da_chot, duong_dan)
                self.m_vua_ky = m
                self.tin_hieu.hoan_thanh_ky.emit(json.dumps({"r": str(r), "s": str(s)}))
            except Exception as e:
                self.tin_hieu.bao_loi.emit("Lỗi Ký", str(e), "ky")

        threading.Thread(target=run_thread, daemon=True).start()

    def _phan_hoi_giao_dien_ky(self, json_str):
        self.o_chu_chu_ky.setText(json_str)
        self.nut_thuc_hien_ky.setEnabled(True)
        QMessageBox.information(self, "Thành công", "Đã tạo chữ ký số!")

    def hanh_dong_xac_thuc(self):
        duong_dan = self.o_nhap_tep_xac_thuc.text().strip()
        chuoi_ck = self.o_chu_chu_ky_xac_thuc.toPlainText().strip()
        if not duong_dan or not chuoi_ck: return

        self.nut_thkh_xac_thuc.setEnabled(False)

        def run_thread():
            try:
                data = json.loads(chuoi_ck)
                m_goc = getattr(self, "m_vua_ky", 0)
                trang_thai = core.xac_thuc_chu_ky_chi_tiet(
                    self.q, self.a, self.YA, int(data["r"]), int(data["s"]), int(m_goc), duong_dan
                )
                self.tin_hieu.hoan_thanh_xac_thuc.emit(trang_thai)
            except Exception as e:
                self.tin_hieu.bao_loi.emit("Lỗi", f"Chữ ký sai định dạng hoặc không thể phân tích: {e}", "xac_thuc")

        threading.Thread(target=run_thread, daemon=True).start()

    def _phan_hoi_xac_thuc_chi_tiet_ui(self, trang_thai):
        self.nut_thkh_xac_thuc.setEnabled(True)
        if trang_thai == "HOP_LE":
            QMessageBox.information(self, "✅ CHỮ KÝ HỢP LỆ", "Văn bản toàn vẹn và chữ ký chính xác.")
        elif trang_thai == "TEP_DA_DOI":
            QMessageBox.critical(self, "❌ VĂN BẢN KHÔNG TOÀN VẸN",
                                 "Văn bản đã bị sửa đổi so với lúc ký (Mã băm không khớp).")
        elif trang_thai == "CHU_KY_DA_DOI":
            QMessageBox.critical(self, "❌ CHỮ KÝ BỊ SỬA ĐỔI",
                                 "Văn bản không bị thay đổi, nhưng thông số chữ ký (r, s) đã bị chỉnh sửa hoặc làm giả.")
        else:
            QMessageBox.critical(self, "❌ VĂN BẢN VÀ CHỮ KÝ ĐỀU BỊ SỬA ĐỔI",
                                 "Văn bản không toàn vẹn và bản thân chữ ký cũng không khớp.")